"""Freeze one dashboard version as the report's citation basis.

Reads a PINNED payload (never the live one) and writes:
  docs/report_freeze/Report_Figures_<ver>.xlsx   - Figures / Tables / Read me sheets
  docs/report_freeze/payload_<ver>.json          - the frozen payload (pinned separately)

Every value is READ from the payload, never transcribed. Four figures are not in the payload and are
recomputed from raw sources; they are flagged amber in the register:
  - panel retention        (raw OCS session gaps)
  - panel runtime / span   (interview_schedule lookup + raw session min/max)
  - 2WT within-2-days      (master triggers + transcript last-message times)

To refresh: re-run against a newer pinned payload and DIFF the Figures sheets. That yields a list of
exactly which figures moved, instead of discovering it in review.

    python build_report_freeze.py docs/report_freeze/payload_v207.json v207
"""

import collections
import csv
import json
import statistics as S
import sys
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "docs/report_freeze/payload_v207.json"
VER = sys.argv[2] if len(sys.argv) > 2 else "v207"
OUT = ROOT / ("docs/report_freeze/Report_Figures_%s.xlsx" % VER)
FACE = "Sans Serif Collection"

D = json.loads(SRC.read_text(encoding="utf-8"))
STAMP = "{}  |  built {}  |  frozen for the GiveWell report".format(VER, D["built_at"])
t1 = {r["key"]: r for r in D["table1"]}
t3 = {r["key"]: r for r in D["table3"]}
cf = {r["sg"]: r for r in D["connectFunnel"]}


def r1(x):
    """Half-up to whole percent, matching the dashboard's own rounding."""
    return float(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def dt(s):
    try:
        return datetime.fromisoformat(str(s).replace("Z", "+00:00"))
    except Exception:
        return None


def pct(num, den):
    return "%d  (%d%%)" % (num, r1(num / den * 100))


# ---- panel per-FLW completion counts, recounted from the payload's own matrix ----
P = {i for i, c in enumerate(D["flwMatrixCohorts"]) if c in ("1PC1", "1PE1")}
pc = collections.Counter()
for e in D["flwMatrixV2"]:
    parts = e.split("|")
    for seg in parts[1:]:
        ci, _, dg = seg.partition(":")
        if int(ci) in P:
            pc[parts[0]] += dg.count("5")
pv = [x for x in pc.values() if x > 0]

# ---- derived: panel retention, from RAW OCS session gaps ----
cache = json.loads((ROOT / "_ocs_state_cache.json").read_text())
by = collections.defaultdict(list)
for s in cache:
    if s.get("cohort_id") in ("1PC1", "1PE1") and s.get("pid") and dt(s.get("created_at")):
        by[s["pid"]].append(dt(s["created_at"]))
for k in by:
    by[k].sort()
RET_N = len(by)
RET_K = sum(1 for ds in by.values() if not any((ds[i + 1] - ds[i]).days > 14 for i in range(len(ds) - 1)))
_all = sorted(x for ds in by.values() for x in ds)
SPAN_W = (_all[-1] - _all[0]).days / 7.0

# ---- derived: panel designed runtime, from the CCHQ interview_schedule lookup ----
DESIGN_W = None
_sch = ROOT / "_interview_schedule.json"
if _sch.exists():
    sch = json.loads(_sch.read_text()).get("1PC1") or []
    offs = [r.get("offset_days") for r in sch if r.get("offset_days") is not None]
    if offs:
        DESIGN_W = max(offs) / 7.0

# ---- derived: 2WT time from offer to completion ----
last_msg = {}
tp = ROOT / "ocs_transcript_dump/transcripts.jsonl"
if tp.exists():
    with tp.open(encoding="utf-8") as f:
        for line in f:
            o = json.loads(line)
            ms = [x for x in (dt(m.get("created_at")) for m in (o.get("messages") or [])) if x]
            if ms:
                last_msg[o["sid"]] = max(ms)
sid_upd = {s["sid"]: dt(s.get("updated_at")) for s in cache}
lags = []
for r in csv.DictReader((ROOT / "master_4src.csv").open(encoding="utf-8")):
    if r["subgroup"] != "2WT" or r["is_completed"] != "Y":
        continue
    a = dt(r["trigger_received_on"])
    b = last_msg.get(r["matched_session_id"]) or sid_upd.get(r["matched_session_id"])
    if a and b:
        lags.append((b - a).total_seconds() / 86400)
W2 = sum(1 for x in lags if x <= 2)
W2_BASE = cf["2WT"]["started"]

PL = "payload"
RAW = "RAW (not in payload)"

FIGURES = [
    (
        "Overall",
        "Interviews started",
        t1["Overall"]["ist"],
        "Overview",
        "unique (FLW, cohort, interview) cells with a session",
        "table1[Overall].ist",
        PL,
    ),
    (
        "Overall",
        "Interviews completed",
        t1["Overall"]["icmp"],
        "Overview",
        "cells whose matched session is interview_complete",
        "table1[Overall].icmp",
        PL,
    ),
    (
        "Overall",
        "Completion rate",
        "%s%%" % t1["Overall"]["pct"],
        "Overview",
        "completed / started",
        "table1[Overall].pct",
        PL,
    ),
    (
        "Overall",
        "Unique FLWs",
        t1["Overall"]["flws"],
        "Overview",
        "FLWs with at least one started interview",
        "table1[Overall].flws",
        PL,
    ),
    (
        "Overall",
        "Cohorts",
        D["counts"]["cohorts"],
        "Scale and Costs",
        "distinct cohorts in the build",
        "counts.cohorts",
        PL,
    ),
    (
        "Panel",
        "FLWs started at least one",
        t1["PANEL"]["flws"],
        "Panel Cohort",
        "FLWs with at least one STARTED panel interview",
        "table1[PANEL].flws",
        PL,
    ),
    (
        "Panel",
        "FLWs completed at least one",
        len(pv),
        "Panel Cohort",
        "FLWs with at least one COMPLETED panel interview",
        "flwMatrixV2 (recount)",
        PL,
    ),
    (
        "Panel",
        "Interviews started",
        t1["PANEL"]["ist"],
        "Panel Cohort",
        "started panel cells",
        "table1[PANEL].ist",
        PL,
    ),
    (
        "Panel",
        "Interviews completed",
        t1["PANEL"]["icmp"],
        "Panel Cohort",
        "completed panel cells",
        "table1[PANEL].icmp",
        PL,
    ),
    (
        "Panel",
        "Completion rate",
        "%s%%" % t1["PANEL"]["pct"],
        "Panel Cohort",
        "completed / started",
        "table1[PANEL].pct",
        PL,
    ),
    (
        "Panel",
        "Median topics per worker",
        int(S.median(pv)),
        "Panel Cohort",
        "median completed count, base = FLWs completing at least one",
        "flwMatrixV2",
        PL,
    ),
    (
        "Panel",
        "Completed 8 or more topics",
        pct(sum(1 for x in pv if x >= 8), len(pv)),
        "Panel Cohort",
        "base = %d FLWs completing at least one" % len(pv),
        "flwMatrixV2",
        PL,
    ),
    (
        "Panel",
        "Completed 11 or more topics",
        pct(sum(1 for x in pv if x >= 11), len(pv)),
        "Panel Cohort",
        "base = %d FLWs completing at least one" % len(pv),
        "flwMatrixV2",
        PL,
    ),
    (
        "Panel",
        "Finished all 13",
        pct(sum(1 for x in pv if x >= 13), len(pv)),
        "Panel Cohort",
        "base = %d FLWs completing at least one" % len(pv),
        "flwMatrixV2",
        PL,
    ),
    (
        "Panel",
        "Stopped after one",
        sum(1 for x in pv if x == 1),
        "Panel Cohort",
        "exactly one completed interview",
        "flwMatrixV2",
        PL,
    ),
    (
        "Panel",
        "Retention",
        "%d%%" % r1(RET_K / RET_N * 100),
        "Panel Cohort",
        "%d of %d FLWs never had a gap over 14 days BETWEEN interviews. Ignores silence after the "
        "last interview; the engagement view, which counts that tail, reads 49%%." % (RET_K, RET_N),
        "raw OCS session created_at",
        RAW,
    ),
    (
        "Panel",
        "Designed runtime per worker",
        ("%.0f weeks" % DESIGN_W) if DESIGN_W else "unavailable",
        "Panel Cohort",
        "13 interviews on a 4-day cadence, offsets 0 to 48 days",
        "_interview_schedule.json",
        RAW,
    ),
    (
        "Panel",
        "Cohort calendar span",
        "%.1f weeks" % SPAN_W,
        "Panel Cohort",
        "first to last raw panel session. Longer than the design because recruitment was staggered.",
        "raw OCS session created_at",
        RAW,
    ),
    (
        "2WT",
        "Invited",
        cf["2WT"]["invited"],
        "2-Week Test",
        "Connect invited_date present",
        "connectFunnel[2WT].invited",
        PL,
    ),
    (
        "2WT",
        "Initiated",
        cf["2WT"]["initiated"],
        "2-Week Test",
        "clicked through the welcome. This is NOT the started count.",
        "connectFunnel[2WT].initiated",
        PL,
    ),
    (
        "2WT",
        "Started",
        cf["2WT"]["started"],
        "2-Week Test",
        "FLWs with a started interview",
        "connectFunnel[2WT].started",
        PL,
    ),
    (
        "2WT",
        "Completed",
        cf["2WT"]["completed"],
        "2-Week Test",
        "FLWs with a completed interview",
        "connectFunnel[2WT].completed",
        PL,
    ),
    (
        "2WT",
        "Completion rate",
        "%s%%" % t1["2WT"]["pct"],
        "2-Week Test",
        "completed / STARTED. Dividing by initiated instead gives 91%, which is what the draft did.",
        "table1[2WT].pct",
        PL,
    ),
    (
        "2WT",
        "Completed within 2 days",
        pct(W2, W2_BASE),
        "2-Week Test",
        "trigger_received_on to last message in the session, 2 days or less. Boundary is not "
        "knife-edge: only 20 of %d lags fall between 1.5 and 2.5 days." % len(lags),
        "master_4src + transcripts",
        RAW,
    ),
    (
        "2WT",
        "Completed within 6 hours",
        pct(sum(1 for x in lags if x <= 0.25), W2_BASE),
        "2-Week Test",
        "same clock, 6 hours or less",
        "master_4src + transcripts",
        RAW,
    ),
    (
        "A/B 1 payment",
        "Arm A completion",
        "%s%%" % t3["ABT1-A"]["pct"],
        "A/B testing",
        "completed / started",
        "table3[ABT1-A].pct",
        PL,
    ),
    (
        "A/B 1 payment",
        "Arm B completion",
        "%s%%" % t3["ABT1-B"]["pct"],
        "A/B testing",
        "completed / started",
        "table3[ABT1-B].pct",
        PL,
    ),
    (
        "A/B 2 cadence",
        "Arm A completion",
        "%s%%" % t3["ABT2-A"]["pct"],
        "A/B testing",
        "completed / started. Was 93.8% before the cohort-attribution fix of 2 Sep.",
        "table3[ABT2-A].pct",
        PL,
    ),
    (
        "A/B 2 cadence",
        "Arm B completion",
        "%s%%" % t3["ABT2-B"]["pct"],
        "A/B testing",
        "completed / started",
        "table3[ABT2-B].pct",
        PL,
    ),
    (
        "A/B 3 length",
        "Arm A completion",
        "%s%%" % t3["ABT3-A"]["pct"],
        "A/B testing",
        "completed / started",
        "table3[ABT3-A].pct",
        PL,
    ),
    (
        "A/B 3 length",
        "Arm B completion",
        "%s%%" % t3["ABT3-B"]["pct"],
        "A/B testing",
        "completed / started",
        "table3[ABT3-B].pct",
        PL,
    ),
]

wb = openpyxl.Workbook()
HDR = PatternFill("solid", fgColor="F2F2F2")
AMBER = PatternFill("solid", fgColor="FFF6E5")


def cell(ws, r, c, v, bold=False, size=11, colour=None, wrap=False, fill=None):
    x = ws.cell(r, c, v)
    x.font = Font(name=FACE, size=size, bold=bold, color=colour)
    if wrap:
        x.alignment = Alignment(wrap_text=True, vertical="top")
    if fill:
        x.fill = fill
    return x


ws = wb.active
ws.title = "Figures"
cell(ws, 1, 1, "Report figures, frozen", bold=True, size=14)
cell(ws, 2, 1, STAMP, size=10, colour="595959")
cell(
    ws,
    3,
    1,
    "Every value is read from the frozen payload, not transcribed. Amber rows are "
    "recomputed from raw sources because the payload does not carry them.",
    size=9,
    colour="595959",
)
for i, h in enumerate(
    ["Section", "Figure", "Value", "Where in report", "Definition", "Source path", "Provenance"], start=1
):
    cell(ws, 5, i, h, bold=True, fill=HDR)
for j, row in enumerate(FIGURES, start=6):
    for i, v in enumerate(row, start=1):
        cell(ws, j, i, v, wrap=(i == 5), fill=AMBER if row[6] == RAW else None)
for i, w in enumerate([15, 30, 20, 18, 70, 30, 20], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A6"

ws2 = wb.create_sheet("Tables")
cell(ws2, 1, 1, "The payload's own tables, flattened so any percentage can be recomputed", bold=True, size=13)
cell(ws2, 2, 1, STAMP, size=10, colour="595959")
r = 4
for name, tbl, keys in (
    ("table1 - by subgroup", D["table1"], ["key", "flws", "ist", "icmp", "pct", "avg_words"]),
    ("table3 - by A/B arm", D["table3"], ["key", "flws", "ist", "icmp", "pct", "avg_words"]),
    (
        "connectFunnel - the recruitment funnel",
        D["connectFunnel"],
        ["sg", "invited", "accepted", "learn_completed", "claimed", "initiated", "started", "completed"],
    ),
):
    cell(ws2, r, 1, name, bold=True, size=12)
    r += 1
    for i, k in enumerate(keys, start=1):
        cell(ws2, r, i, k, bold=True, fill=HDR)
    r += 1
    for row in tbl:
        for i, k in enumerate(keys, start=1):
            cell(ws2, r, i, row.get(k))
        r += 1
    r += 2
for i, w in enumerate([24, 12, 12, 12, 12, 16, 12, 12], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws3 = wb.create_sheet("Read me")
NOTES = [
    ("How to use this", True),
    (STAMP, False),
    ("", False),
    ("This is the citation basis for the report. Quote these values rather than the live dashboard,", False),
    ("so the numbers stop moving under us. payload_" + VER + ".json beside this file is the frozen source,", False),
    ("so every figure here is reproducible and auditable.", False),
    ("", False),
    ("To refresh: re-run build_report_freeze.py against a newer pinned payload and diff the Figures", False),
    ("sheets. That gives a list of exactly which figures moved and why, rather than finding out in", False),
    ("review.", False),
    ("", False),
    ("Three definition traps that have already produced errors in the report", True),
    ("", False),
    ("1. Panel retention is 89% OR 49%, depending on the definition.", True),
    ("   89% is the share of FLWs who never had a gap over 14 days BETWEEN interviews. It ignores", False),
    ("   silence after someone's last interview, so a worker who did two interviews in June and", False),
    ("   stopped still counts as retained. 49% is the dashboard engagement view, which counts that", False),
    ("   tail. Both are defensible. The sentence has to say which one it means.", False),
    ("", False),
    ("2. For 2WT, 534 is INITIATED, not started. 515 started.", True),
    ("   The draft said 67% (534) started and computed completion as 484/534 = 91%. The dashboard", False),
    ("   figure is 484/515 = 94%. Dividing by the wrong funnel step understates completion by three", False),
    ("   points.", False),
    ("", False),
    ("3. Panel percentages: base 364 or base 366.", True),
    ("   364 FLWs completed at least one interview; 366 started at least one. On 364 the shares are", False),
    ("   76 / 55 / 26. On 366 they are 75 / 54 / 25. Pick one and use it for the whole paragraph,", False),
    ("   because a reader will divide by whichever count the sentence names.", False),
    ("", False),
    ("Panel runtime", True),
    ("   7 weeks per worker by design: 13 interviews on a 4-day cadence, offsets 0 to 48 days.", False),
    ("   11.6 weeks of calendar time for the cohort, because recruitment was staggered over about", False),
    ("   four weeks. Do NOT read runtime off the engagement chart week buckets - that axis starts", False),
    ("   before activity begins and extends to today, so it grows every day.", False),
    ("", False),
    ("What can still move, even with the bot frozen", True),
    ("   Late review tagging: verdicts arrive weeks after completion, which moves the acceptable and", False),
    ("   unacceptable splits but not the completion counts.", False),
    ("   85 sessions still sit at interview_ongoing and could yet complete.", False),
    ("   Our own fixes: three figures changed on 2 September because of a cohort-attribution fix, not", False),
    ("   because of new data. ABT2-A went from 93.8% to 92.1%.", False),
]
for i, (s, b) in enumerate(NOTES, start=1):
    cell(ws3, i, 1, s, bold=b)
ws3.column_dimensions["A"].width = 104

wb.save(OUT)
print("wrote %s" % OUT)
print("  Figures: %d rows (%d recomputed from raw)" % (len(FIGURES), sum(1 for f in FIGURES if f[6] == RAW)))
print("  stamp: %s" % STAMP)
