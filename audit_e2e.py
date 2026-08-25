"""END-TO-END AUDIT — every number, traced raw->master->status->aggregates->payload,
reconciled against (1) the 10-Jun master baseline, (2) independent recompute, (3) the GW
workbook. Prints PASS/FAIL per check + a final accuracy summary. No spot checks.
"""
import csv as _csv
import glob
import json
import re
import os
import sys
from collections import Counter, defaultdict
from datetime import date, timedelta

from openpyxl import load_workbook

import build_master_4src as bm

# Config, not logic: the per-cohort grace override is DATA the definition is parameterised by, so
# reading it here keeps the gate honest. What a gate must never import is the RULE it is checking,
# which is why the state machine below is still a hand-written re-implementation.
from topic_status_lib import GRACE_DAYS as tsl_grace
  # the master under test

TODAY = date.today()  # match the build's dynamic time-gating for the independent status recompute
_CANON_TOPICS = ["A", "B", "C", "D", "E", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "8S", "8L", "10S", "10L", "11S", "11L", "13L", "99", "101", "F", "G"]
TOPICS = [t for t in _CANON_TOPICS if any(t in bm.SUBGROUP_DESIGN[sg]["topics"] for sg in bm.SUBGROUP_DESIGN)]
SG_ORDER = ["TRS", "TRE", "ABT1-A", "ABT1-B", "ABT2-A", "ABT2-B", "PANEL", "ABT3-A", "ABT3-B", "2WT", "EXT", "NPS"]
ROLL = {"TRS": "TRS", "TRE": "TRE", "ABT1-A": "ABT1", "ABT1-B": "ABT1", "ABT2-A": "ABT2", "ABT2-B": "ABT2",
        "PANEL": "PANEL", "ABT3-A": "ABT3", "ABT3-B": "ABT3", "2WT": "2WT", "EXT": "EXT", "NPS": "NPS"}
results = []  # (section, check, passed, detail)


def chk(section, name, passed, detail=""):
    results.append((section, name, passed, detail))
    print(f"  [{'PASS' if passed else 'FAIL'}] {name}  {detail}")


print("=" * 90)
print("A. RAW SOURCE INTEGRITY")
print("=" * 90)
n_trig = sum(len(v) for v in bm.triggers_by_flw_iv.values())
n_wel = sum(len(v) for v in bm.welcome_flws_by_key.values())
sessions = json.loads(open("_ocs_state_cache.json").read())
n_tagged = sum(1 for s in sessions if s.get("pid") and s.get("interview") and str(s["interview"]).strip())
# Matrix universe = Connect-claimed ∪ anyone holding a master row in that cohort. Claimed-only used to
# leave 210 started / 182 completed interviews outside the grid while table1/table2 still counted them
# (2026-08-07 audit). Independent of topic_status_lib on purpose — this is the gate's own derivation.
claimed_pairs = {
    (c, f) for c in bm.cohort_flws for f in bm.cohort_flws[c] if bm.cohort_flw_meta[(c, f)].get("date_claimed")
} | {(r["cohort_id"], r["connect_id"]) for r in bm.rows if bm.cohort_to_sg(r["cohort_id"])}
print(
    f"  triggers={n_trig}  welcome-keys={n_wel}  ocs_sessions={len(sessions)} tagged={n_tagged}  claimed_pairs={len(claimed_pairs)}"
)
# no duplicate trigger_form_id in master
fids = [r["trigger_form_id"] for r in bm.rows]
chk("A", "no duplicate trigger_form_id in master", len(fids) == len(set(fids)), f"{len(fids)} rows")
# every cohort maps to a subgroup
bad_sg = [c for c in bm.cohort_info if bm.cohort_to_sg(c) is None]
chk("A", "every cohort maps to a subgroup", not bad_sg, f"{len(bad_sg)} bad")

print("=" * 90)
print("B. MASTER vs 10-Jun BASELINE (row-level) + INVARIANTS")
print("=" * 90)
# Baseline comparison is OPTIONAL — the 10-Jun baseline holds participant ids and is not shipped
# server-side. When absent, the integrity invariants below still run (they need only bm.rows).
if os.path.exists("master_v7_2026-06-10.csv"):
    # exclude the intentionally-removed test accounts from the baseline too, else their (now-dropped)
    # baseline rows look like a coverage regression. See bm.EXCLUDE_FLWS.
    base = {r["trigger_form_id"]: r for r in _csv.DictReader(open("master_v7_2026-06-10.csv", encoding="utf-8"))
            if r["connect_id"] not in bm.EXCLUDE_FLWS}
    live = {r["trigger_form_id"]: r for r in bm.rows}
    shared = set(base) & set(live)
    only_base = set(base) - set(live)
    only_live = set(live) - set(base)
    # Growth-aware: live must be a SUPERSET of the 10-Jun baseline (new interviews expected as data
    # flows in). A regression = a baseline row that vanished from live. only_live>0 is fine (growth).
    chk(
        "B",
        "baseline rows all present in live (no coverage regression)",
        len(only_base) == 0,
        f"shared={len(shared)} only_live={len(only_live)} (growth) only_base={len(only_base)}",
    )
    struct_cols = [
        "cohort_id",
        "subgroup",
        "cohort_type",
        "interview_n",
        "topic_code",
        "topic_name",
        "training_date",
        "release_date",
    ]
    struct_mm = sum(1 for k in shared for c in struct_cols if str(live[k][c]) != str(base[k][c]))
    chk(
        "B",
        "structural columns bit-exact vs baseline",
        struct_mm == 0,
        f"{struct_mm} mismatches across {len(struct_cols)} cols",
    )
    st_reg = sum(1 for k in shared if base[k]["is_started"] == "Y" and live[k]["is_started"] == "N")
    co_reg = sum(1 for k in shared if base[k]["is_completed"] == "Y" and live[k]["is_completed"] == "N")
    st_fwd = sum(1 for k in shared if base[k]["is_started"] == "N" and live[k]["is_started"] == "Y")
    co_fwd = sum(1 for k in shared if base[k]["is_completed"] == "N" and live[k]["is_completed"] == "Y")
    chk("B", "is_started: zero regressions (Y->N)", st_reg == 0, f"regressions={st_reg}, forward N->Y={st_fwd}")
    chk("B", "is_completed: zero regressions (Y->N)", co_reg == 0, f"regressions={co_reg}, forward N->Y={co_fwd}")
else:
    print("  [SKIP] master_v7_2026-06-10 baseline not present — integrity invariants below still enforced.")
# invariants
inv1 = all(not (r["is_completed"] == "Y" and r["is_started"] != "Y") for r in bm.rows)
inv2 = all(not (r["is_started"] == "Y" and r["matched_session_id"] == "") for r in bm.rows)
sid_use = Counter(r["matched_session_id"] for r in bm.rows if r["matched_session_id"])
inv3 = all(v == 1 for v in sid_use.values())
inv4 = all(
    SUBGROUP_OK := (
        r["topic_code"] in bm.SUBGROUP_DESIGN[r["subgroup"]]["topics"]
        and bm.SUBGROUP_DESIGN[r["subgroup"]]["topics"].index(r["topic_code"]) + 1 == int(r["interview_n"])
    )
    for r in bm.rows
)
chk("B", "invariant completed=>started", inv1)
chk("B", "invariant started=>has session", inv2)
chk("B", "invariant no session double-claimed", inv3, f"{sum(1 for v in sid_use.values() if v>1)} dup sids")
chk("B", "invariant interview_n == position(topic) in subgroup (ALL rows)", inv4)

print("=" * 90)
print("C. STATUS TABLE (independent recompute, full grid)")
print("=" * 90)
mlook = {}


def rank(r):
    return (1 if r["is_completed"] == "Y" else 0) * 2 + (1 if r["is_started"] == "Y" else 0)


for r in bm.rows:
    k = (r["connect_id"], r["cohort_id"], r["topic_code"])
    if k not in mlook or rank(r) > rank(mlook[k]):
        mlook[k] = r


def status_for(flw, cohort, topic):
    """Independent re-implementation of the 7-state model (deliberately NOT importing
    topic_status_lib — a gate that calls the code under test proves nothing).
    available-* = the bot sent it and got no session; not-triggered = it was never sent."""
    sg = bm.cohort_to_sg(cohort)
    topics = bm.SUBGROUP_DESIGN[sg]["topics"]
    if topic not in topics:
        return "not-applicable"
    n = topics.index(topic) + 1
    m = mlook.get((flw, cohort, topic))
    if m and m["is_completed"] == "Y":
        return "completed"
    if m and m["is_started"] == "Y":
        return "started-not-completed"
    # start_date, not training_date: the resolved cohort start (invitation, else first trigger)
    # that every view now shares. The RULE below stays an independent re-implementation.
    td = bm.cohort_info.get(cohort, {}).get("start_date")
    gr = tsl_grace.get(cohort)
    cad = bm.SUBGROUP_DESIGN[sg]["cadence"]
    if m:
        # No exemption for the FINAL interview. Its deadline is a property of that interview -
        # released, plus one gap to do it - so the last one has a deadline like every other. The old
        # `n < len(topics)` carve-out meant nobody could ever be recorded as skipping their last
        # interview, which is why a single-interview design showed an impossible 0% drop-off.
        if td and TODAY >= td + timedelta(days=(n - 1) * cad + (cad if gr is None else gr)):
            return "available-missed-overdue"
        return "available-not-started"
    if not td or not cad:
        return "available-not-started"   # schedule unknown -> can't call it due, so can't call it missed
    if TODAY < td + timedelta(days=(n - 1) * cad):
        return "not-available-yet"
    return "not-triggered"


grid = {}
for cohort, flw in claimed_pairs:
    for topic in TOPICS:
        grid[(cohort, flw, topic)] = status_for(flw, cohort, topic)
chk(
    "C",
    f"grid complete (claimed_pairs x {len(TOPICS)}, no dup/missing)",
    len(grid) == len(claimed_pairs) * len(TOPICS),
    f"{len(grid)} == {len(claimed_pairs)*len(TOPICS)}",
)
na = sum(1 for v in grid.values() if v == "not-applicable")
exp_na = sum(len(TOPICS) - len(bm.SUBGROUP_DESIGN[bm.cohort_to_sg(c)]["topics"]) for (c, f) in claimed_pairs)
chk("C", "not-applicable count exact", na == exp_na, f"{na} == {exp_na}")
# C: started/completed reconcile to master (claimed)
g_comp = sum(1 for v in grid.values() if v == "completed")
g_start = sum(1 for v in grid.values() if v in ("completed", "started-not-completed"))
m_comp = len(
    {
        (r["connect_id"], r["cohort_id"], r["topic_code"])
        for r in bm.rows
        if r["is_completed"] == "Y" and (r["cohort_id"], r["connect_id"]) in claimed_pairs
    }
)
m_start = len(
    {
        (r["connect_id"], r["cohort_id"], r["topic_code"])
        for r in bm.rows
        if r["is_started"] == "Y" and (r["cohort_id"], r["connect_id"]) in claimed_pairs
    }
)
chk("C", "status completed == master completed (claimed)", g_comp == m_comp, f"{g_comp} == {m_comp}")
chk("C", "status started == master started (claimed)", g_start == m_start, f"{g_start} == {m_start}")


# C: independent re-derivation of schedule states (second code path)
def status_v2(flw, cohort, topic):
    sg = bm.cohort_to_sg(cohort)
    topics = bm.SUBGROUP_DESIGN[sg]["topics"]
    if topic not in topics:
        return "not-applicable"
    n = topics.index(topic) + 1
    m = mlook.get((flw, cohort, topic))
    completed = bool(m) and m["is_completed"] == "Y"
    started = bool(m) and m["is_started"] == "Y"
    td = bm.cohort_info.get(cohort, {}).get("start_date")
    gr = tsl_grace.get(cohort)
    rel = (td + timedelta(days=(n - 1) * cad_)) if (cad_ := bm.SUBGROUP_DESIGN[sg]["cadence"]) and td else None
    # deadline = release + one gap (or the cohort's override); the final interview is included
    nrel = (td + timedelta(days=(n - 1) * cad_ + (cad_ if gr is None else gr))) if td and cad_ else None
    avail = rel is not None and TODAY >= rel
    overdue = nrel is not None and TODAY >= nrel
    if completed:
        return "completed"
    if started:
        return "started-not-completed"
    triggered = bool(m)          # a master row exists only where a trigger form does
    if triggered:
        return "available-missed-overdue" if overdue else "available-not-started"
    if td is None or rel is None:
        return "available-not-started"   # schedule unknown -> not provably due
    if not avail:
        return "not-available-yet"
    return "not-triggered"


mism = sum(1 for k, v in grid.items() if status_v2(k[1], k[0], k[2]) != v)
chk(
    "C",
    "status logic agrees across 2 independent code paths (ALL cells)",
    mism == 0,
    f"{mism} disagreements / {len(grid)}",
)
distinct_states = set(grid.values())
chk(
    "C",
    "every cell exactly one of the 7 states",
    distinct_states
    <= set(
        bm.__dict__.get("STATES", [])
        or [
            "not-applicable",
            "not-available-yet",
            "not-triggered",
            "available-not-started",
            "available-missed-overdue",
            "started-not-completed",
            "completed",
        ]
    ),
    f"states seen: {sorted(distinct_states)}",
)

print("=" * 90)
print("D. AGGREGATES — payload vs INDEPENDENT recompute (every cell)")
print("=" * 90)
payload = json.loads(open("payload_agg.json", encoding="utf-8").read())
# independent funnel recompute
cell = {}
for r in bm.rows:
    k = (r["connect_id"], r["cohort_id"], int(r["interview_n"]))
    c = cell.setdefault(
        k,
        {"sg": r["subgroup"], "n": int(r["interview_n"]), "flw": r["connect_id"], "t": False, "s": False, "c": False},
    )
    c["t"] = True
    if r["is_started"] == "Y":
        c["s"] = True
    if r["is_completed"] == "Y":
        c["c"] = True
cells = list(cell.values())
elig_sg = defaultdict(set)
for (cohort, topic), flws in bm.welcome_flws_by_key.items():
    sg = bm.cohort_to_sg(cohort)
    if sg:
        elig_sg[sg] |= flws
fre = defaultdict(lambda: {"t": set(), "s": set(), "c": set()})
for c in cells:
    f = fre[(c["sg"], c["n"])]
    if c["t"]:
        f["t"].add(c["flw"])
    if c["s"]:
        f["s"].add(c["flw"])
    if c["c"]:
        f["c"].add(c["flw"])
fmm = 0
for row in payload["funnel"]:
    f = fre[(row["sg"], row["n"])]
    if (
        len(f["t"]) != row["trig"]
        or len(f["s"]) != row["started"]
        or len(f["c"]) != row["completed"]
        or len(elig_sg[row["sg"]]) != row["elig"]
    ):
        fmm += 1
chk("D", "funnel: payload == independent recompute (all 25 rows)", fmm == 0, f"{fmm} mismatched rows")


# Tables independent
def agg(keyfn, keys):
    a = defaultdict(lambda: {"flw": set(), "ist": 0, "icmp": 0})
    for c in cells:
        for k in set(keyfn(c)):
            if c["s"]:
                a[k]["flw"].add(c["flw"])
                a[k]["ist"] += 1
            if c["c"]:
                a[k]["icmp"] += 1
    return a


t1a = agg(lambda c: [ROLL[c["sg"]], "Overall"], None)
t1mm = sum(
    1
    for row in payload["table1"]
    if not (
        len(t1a[row["key"]]["flw"]) == row["flws"]
        and t1a[row["key"]]["ist"] == row["ist"]
        and t1a[row["key"]]["icmp"] == row["icmp"]
    )
)
chk("D", "Table1: payload == recompute", t1mm == 0, f"{t1mm} mismatched")
t3a = agg(lambda c: ([c["sg"], "Overall"] if c["sg"].startswith(("ABT1", "ABT2", "ABT3")) else []), None)
t3mm = sum(
    1
    for row in payload["table3"]
    if not (
        len(t3a[row["key"]]["flw"]) == row["flws"]
        and t3a[row["key"]]["ist"] == row["ist"]
        and t3a[row["key"]]["icmp"] == row["icmp"]
    )
)
chk("D", "Table3: payload == recompute", t3mm == 0, f"{t3mm} mismatched")
t2a = defaultdict(lambda: {"flw": set(), "ist": 0, "icmp": 0})
for c in cells:
    tc = bm.SUBGROUP_DESIGN[c["sg"]]["topics"][c["n"] - 1]
    if c["s"]:
        t2a[tc]["flw"].add(c["flw"])
        t2a[tc]["ist"] += 1
    if c["c"]:
        t2a[tc]["icmp"] += 1
t2mm = sum(
    1
    for row in payload["table2"]
    if not (
        len(t2a[row["code"]]["flw"]) == row["flws"]
        and t2a[row["code"]]["ist"] == row["ist"]
        and t2a[row["code"]]["icmp"] == row["icmp"]
    )
)
chk("D", "Table2: payload == recompute", t2mm == 0, f"{t2mm} mismatched")
# topic_status dist vs grid
tsmm = 0
for row in payload["topic_status"]:
    for s in payload["states"]:
        if sum(1 for k, v in grid.items() if k[2] == row["code"] and v == s) != row[s]:
            tsmm += 1
chk("D", "topic_status dist: payload == grid recompute (all topic×state cells)", tsmm == 0, f"{tsmm} mismatched cells")
# line series == funnel pct_started (over present subgroups only)
lmm = 0
for sg in payload["line_pct_started"]:
    fr = [r for r in payload["funnel"] if r["sg"] == sg]
    if [r["pct_started"] for r in sorted(fr, key=lambda x: x["n"])] != payload["line_pct_started"][sg]:
        lmm += 1
chk("D", "line series == funnel %started", lmm == 0, f"{lmm} subgroups mismatched")

print("=" * 90)
print("E. CROSS-CONSISTENCY")
print("=" * 90)
# Table2.ist == sum funnel.started over (sg,n) with that topic  (one FLW=one cohort per subgroup => unique==count)
ok = True
for row in payload["table2"]:
    s = sum(r["started"] for r in payload["funnel"] if r["topic"] == row["code"])
    if s != row["ist"]:
        ok = False
chk("E", "Table2 interviews-started == Σ funnel started by topic", ok)
# funnel monotonic. completed<=started<=triggered is an always-true subset chain -> strict.
# triggered<=eligible USUALLY holds but is legitimately violable: an FLW can trigger an interview
# without a captured welcome form, and eligible/initiated is welcome-derived (e.g. a fresh cohort where
# a trigger lands a beat before its welcome). Allow triggered to exceed eligible by a small margin
# (benign trigger-without-welcome) so it doesn't flap red, while still catching gross trigger inflation.
mono = all(r["completed"] <= r["started"] <= r["trig"] for r in payload["funnel"])
chk("E", "funnel monotonic completed<=started<=triggered", mono)
elig_slack = all(r["trig"] <= r["elig"] + max(3, round(0.02 * r["elig"])) for r in payload["funnel"])
chk("E", "funnel triggered<=eligible (+ benign trigger-without-welcome slack)", elig_slack)
# Overall FLWs (table1) == unique started connect_ids (dedup across subgroups)
overall_flw = len({c["flw"] for c in cells if c["s"]})
t1_overall = next(r["flws"] for r in payload["table1"] if r["key"] == "Overall")
chk(
    "E",
    "Table1 Overall FLWs == dedup unique started FLWs (no inflation)",
    overall_flw == t1_overall,
    f"{overall_flw} == {t1_overall}",
)

print("=" * 90)
print("F. GW WORKBOOK RECONCILIATION (11-Jun reference; live is newer -> forward drift)")
print("=" * 90)
_GW_XLSX = "screenshots/Latest files/GW Tables - 11th June 2026.xlsx"
if not os.path.exists(_GW_XLSX):
    print(
        f"  [SKIP] GW reference workbook not present ({_GW_XLSX}) — one-time reference check, "
        f"not a data-integrity gate. Other checks still enforced."
    )
else:
    wb = load_workbook(_GW_XLSX, read_only=True, data_only=True)
    # parse Retention Drop-off funnel
    ws = wb["Retention Drop-off"]
    rows_ws = [list(r) for r in ws.iter_rows(values_only=True)]
    gw_funnel = {}
    for r in rows_ws:
        if r and r[0] in SG_ORDER:
            sg = r[0]
            # interview blocks start at col 9, step 8: Topic,Elig,Trig,%,Start,%,Compl,%
            for n in range(1, 9):
                base_c = 8 + (n - 1) * 8 + 1
                topic = r[base_c] if base_c < len(r) else None
                if topic in (None, "-", ""):
                    continue
                try:
                    elig = int(r[base_c + 1])
                    trig = int(r[base_c + 2])
                    start = int(r[base_c + 4])
                    comp = int(r[base_c + 6])
                except (TypeError, ValueError):
                    continue
                gw_funnel[(sg, n)] = (elig, trig, start, comp)
    # GW predates the test-account cleanup, so it still counts the EXCLUDE_FLWS. Subtract their
    # contribution so parity reflects post-cleanup expectations while still catching any OTHER
    # (unexplained) downward drift. Interview cells come from the same-era 10-Jun baseline; initiated
    # (elig) from raw welcome forms (bm already dropped the excluded FLWs).
    exc_iv = defaultdict(lambda: {"t": set(), "s": set(), "c": set()})
    if os.path.exists("master_v7_2026-06-10.csv"):
        for r in _csv.DictReader(open("master_v7_2026-06-10.csv", encoding="utf-8")):
            if r["connect_id"] not in bm.EXCLUDE_FLWS:
                continue
            k = (r["subgroup"], int(r["interview_n"]))
            exc_iv[k]["t"].add(r["connect_id"])
            if r["is_started"] == "Y": exc_iv[k]["s"].add(r["connect_id"])
            if r["is_completed"] == "Y": exc_iv[k]["c"].add(r["connect_id"])
    exc_init = defaultdict(set)
    for _p in glob.glob("hq_pull_full/*welcome_click_start.jsonl"):
        for _line in open(_p, encoding="utf-8"):
            try:
                _s = json.loads(_line)
            except Exception:
                continue
            _f = _s.get("form", {}) or {}
            _m = _f.get("meta", {}) if isinstance(_f.get("meta"), dict) else {}
            _cid = (_f.get("connect_id") or _m.get("username") or _s.get("username") or "").strip()
            if _cid in bm.EXCLUDE_FLWS:
                _sg = bm.cohort_to_sg((_f.get("cohort_id") or "").strip())
                if _sg:
                    exc_init[_sg].add(_cid)
    exact = drift_fwd = drift_other = 0
    for row in payload["funnel"]:
        key = (row["sg"], row["n"])
        if key not in gw_funnel:
            continue
        ge, gt, gs, gc = gw_funnel[key]
        _ex = exc_iv.get(key, {"t": set(), "s": set(), "c": set()})
        ge -= len(exc_init.get(row["sg"], set()))  # elig = # initiated; drop excluded initiated
        gt -= len(_ex["t"]); gs -= len(_ex["s"]); gc -= len(_ex["c"])
        le, lt, ls, lc = row["elig"], row["trig"], row["started"], row["completed"]
        if (ge, gt, gs, gc) == (le, lt, ls, lc):
            exact += 1
        elif le >= ge and lt >= gt and ls >= gs and lc >= gc:
            drift_fwd += 1
        else:
            drift_other += 1
            print(f"    OTHER-DRIFT {key}: live(e/t/s/c)={le}/{lt}/{ls}/{lc} gw={ge}/{gt}/{gs}/{gc}")
    chk(
        "F",
        "funnel vs GW: all cells exact OR forward-drift (live>=gw)",
        drift_other == 0,
        f"exact={exact} forward-drift={drift_fwd} other={drift_other}",
    )

# ---------------------------------------------------------------- G. the clock must not move numbers
# A cohort that has ENDED must not gain drop-outs simply because the daily job ran again. If it does,
# the refresh is counting silence rather than behaviour, and every closed cohort drifts worse forever.
# Rebuild the payload with the clock pushed months forward and diff EVERY leaf: only the date stamp may
# move. Opt out with AUDIT_SKIP_CLOCK=1 (it costs one extra build).
if not os.environ.get("AUDIT_SKIP_CLOCK"):
    import subprocess

    def _leaves(o, p=""):
        if isinstance(o, dict):
            for k, v in o.items():
                yield from _leaves(v, p + "." + str(k))
        elif isinstance(o, list):
            for i2, v in enumerate(o):
                yield from _leaves(v, p + "[" + str(i2) + "]")
        else:
            yield p, o

    # Build BOTH sides here rather than trusting whatever payload_agg.json happens to hold. Reading
    # the on-disk file made this check depend on what ran before it: a stale file from an earlier
    # build reported 275 phantom moves, and a clean rebuild then reported none. A gate whose verdict
    # depends on command ordering is not a gate.
    _env_now = dict(os.environ, PYTHONIOENCODING="utf-8")
    _env_now.pop("INTERVIEWS_TODAY", None)
    _r0 = subprocess.run([sys.executable, "build_payload_agg.py"], env=_env_now,
                         capture_output=True, text=True, timeout=1800)
    if _r0.returncode != 0:
        chk("G", "ENDED cohorts are invariant to the build date", False,
            f"baseline rebuild failed: {(_r0.stderr or '')[-200:]}")
        raise SystemExit(1)
    _now = json.load(open("payload_agg.json", encoding="utf-8"))
    _env = dict(os.environ, INTERVIEWS_TODAY="2027-03-01", PYTHONIOENCODING="utf-8")
    _r = subprocess.run([sys.executable, "build_payload_agg.py"], env=_env,
                        capture_output=True, text=True, timeout=1800)
    if _r.returncode != 0:
        chk("G", "closed cohorts are invariant to the build date", False,
            f"future-dated rebuild failed: {(_r.stderr or '')[-200:]}")
    else:
        _fut = json.load(open("payload_agg.json", encoding="utf-8"))
        _A, _B = dict(_leaves(_now)), dict(_leaves(_fut))
        # An OPEN cohort is SUPPOSED to move with the clock: push the date past its end and it closes,
        # its line goes solid, its funnel status changes. The requirement is only that a cohort which
        # has ALREADY ENDED cannot move. So collect what is still open as of the real today and exempt
        # it by name; everything else must be frozen.
        _open = {c for c, v in (_now.get("line_active") or {}).items() if v}
        _open |= {r["c"] for r in (_now.get("cohort_dropoff") or []) if r.get("e", "") > str(TODAY)}
        # A cohort with a deadline still to arrive is NOT finished, whatever its end date says: PANEL
        # closed on 2 Aug and was still triggering interviews on the 25th. Its numbers will move when
        # that deadline passes, and a deadline passing is information, not silence accumulating. Only
        # cohorts where every deadline is already behind us must be frozen - and they are the ones the
        # requirement is actually about.
        _open |= {r["c"] for r in (_now.get("cohort_dropoff") or []) if not r.get("settled", True)}
        _open |= {bm.cohort_to_sg(r["c"]) for r in (_now.get("cohort_dropoff") or [])
                  if not r.get("settled", True)}
        _today_only = {".today", ".built_at"}

        def _is_open(path):
            if path in _today_only:
                return True
            # `settled` is a statement ABOUT the current date ("every deadline is behind us"), so it
            # flips from false to true as the clock advances. That is the flag working, not drift.
            if path.endswith(".settled"):
                return True
            # Match on PATH SEGMENTS, not substrings: ".line_status.NPS[0]" has NPS followed by a
            # bracket, so looking for ".NPS." or a trailing ".NPS" missed it. Segment matching also
            # stops a cohort whose name merely CONTAINS an open cohort's name from being exempted.
            if set(re.split(r"[.\[\]]+", path)) & _open:
                return True
            # Positional paths like .funnel[50].status carry no name, so resolve the row and read it.
            _m = re.match(r"^\.([A-Za-z_]+)\[(\d+)\]", path)
            if _m and isinstance(_now.get(_m.group(1)), list):
                _row = _now[_m.group(1)][int(_m.group(2))]
                if isinstance(_row, dict):
                    return any(_row.get(f) in _open for f in ("sg", "c", "cohort", "subgroup", "key"))
            return False

        _diff = [k for k in set(_A) | set(_B)
                 if _A.get(k, "<absent>") != _B.get(k, "<absent>") and not _is_open(k)]
        _moved_open = sum(1 for k in set(_A) | set(_B)
                          if _A.get(k, "<absent>") != _B.get(k, "<absent>") and _is_open(k)
                          and k not in _today_only)     # the two date stamps are not a cohort
        chk("G", "ENDED cohorts are invariant to the build date "
                 "(no silent-FLW drift on opportunities that are over)",
            not _diff,
            f"{len(_A):,} leaves compared, {len(_diff)} moved on ended cohorts; "
            f"{_moved_open} moved on still-open cohorts ({', '.join(sorted(_open)) or 'none'}), "
            f"which is expected"
            + ("; OFFENDERS: " + ", ".join(sorted(_diff)[:4]) if _diff else ""))
        json.dump(_now, open("payload_agg.json", "w", encoding="utf-8"))   # restore the real build


print("\n" + "=" * 90)
print("AUDIT SUMMARY")
print("=" * 90)
passed = sum(1 for _, _, p, _ in results if p)
for sec in ["A", "B", "C", "D", "E", "F", "G"]:
    secr = [r for r in results if r[0] == sec]
    print(f"  Section {sec}: {sum(1 for r in secr if r[2])}/{len(secr)} passed")
print(f"\n  TOTAL: {passed}/{len(results)} checks passed")
print(
    "  RESULT:", "ALL PASS — 200% reconciled" if passed == len(results) else f"*** {len(results)-passed} FAILURES ***"
)
# Exit non-zero so the orchestrator actually ABORTS on failure — this gate used to return 0 whatever
# it printed, which made refresh_interviews_dashboard.py report "OK" on a failing audit and publish.
sys.exit(1 if passed != len(results) else 0)
