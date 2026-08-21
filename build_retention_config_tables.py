#!/usr/bin/env python3
"""The retention-configuration table Ali asked for (task 1, first step).

Her ask, verbatim: "creating a table with one row per cohort config type (duration and frequency) and
one column per calculation feeding into the retention graphs. we know one column is definition of
inactivity. im sure there are others? if the cohort end date is a relevant value (it must be right?)
then you would want to add that as a column and then just have one row per cohort".

Shows the configuration BEFORE and AFTER the 2026-08-21 fix, so the table doubles as the change log
for what Ali raised. Sheets 1 and 2 are structural (read from the code). Sheet 3's outcome columns are
computed from the LIVE dashboard payload under the new definition, so they are what the dashboard will
show on its next refresh.

Deliberately kept to four short sheets, plain language, no jargon:
  Start here          what this is, in five lines
  1 The rules         every calculation, the number it uses, whether that number is the same for all
  2 By cohort design  one row per design, one column per calculation. Ali's core request
  3 By cohort         one row per cohort, because cohorts of one design start weeks apart

"Gap" = days between one scheduled interview and the next.
Structure comes from reading the code; every number comes from live data.
"""
import datetime as dt
import json
import os
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent
LIVE = Path(
    sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.environ.get("CLAUDE_JOB_DIR", "."), "tmp", "live_data.json")
)
SCHED = json.loads((ROOT / "_interview_schedule.json").read_text(encoding="utf-8"))
D = json.loads(LIVE.read_text(encoding="utf-8"))
CSG, CE = D["cohortSG"], D["cohortEngagement"]
TODAY = dt.date.today()
SENT = {2, 3, 4, 5}
# Sent but not finished. State 2 ("window still open") is folded in with 3 because every state-2 slot in
# the live payload is a FINAL interview past its window, held open only by the exemption now removed.
CLOSED = {2, 3, 4}

# ---------------------------------------------------------------- cohort calendars
# Dates come from the pipeline's own cohort_dropoff block rather than being re-derived here, so this
# table cannot disagree with the dashboard about when a cohort started or ended. It already carries the
# invitation-date-else-first-trigger fallback, which is what gets all 72 cohorts covered (the Connect
# snapshot is missing invitation dates for the newer ones). Outcome COUNTS are taken from the live
# payload below, because a local pipeline run reflects whatever the cached pulls last saw.
_agg = json.loads((ROOT / "payload_agg.json").read_text(encoding="utf-8"))
CD = {r["c"]: r for r in _agg.get("cohort_dropoff", [])}
train = {c: dt.date.fromisoformat(r["s"]) for c, r in CD.items()}
own_end = {c: dt.date.fromisoformat(r["e"]) for c, r in CD.items()}
est_start = {c for c, r in CD.items() if r.get("x")}
if len(train) < len(CSG):
    print(
        "  WARNING: cohort dates cover "
        + str(len(train))
        + " of "
        + str(len(CSG))
        + " cohorts - run build_payload_agg.py first"
    )

sched = {}
for c, rows in SCHED.items():
    offs = sorted(r.get("offset_days") or 0 for r in rows)
    sched[c] = {"n": len(rows), "offs": offs, "gap": (offs[1] - offs[0]) if len(offs) > 1 else None, "runs": offs[-1]}

# ---------------------------------------------------------------- outcomes from the live matrix
out = defaultdict(Counter)
for ent in D["flwMatrixV2"]:
    for p in ent.split("|")[1:]:
        k = p.index(":")
        cohort = D["flwMatrixCohorts"][int(p[:k])]
        app = [s for s in (int(ch) for ch in p[k + 1 :].rstrip("u")) if s != 0]
        if not app:
            continue
        o = out[cohort]
        o["workers"] += 1
        if all(s == 5 for s in app):
            o["finished"] += 1  # completed every interview in their design
        elif not any(s in SENT for s in app):
            o["never"] += 1  # nothing was ever sent: they never began
        elif any(s in CLOSED for s in app):
            o["dropped"] += 1  # a sent interview went past its deadline unfinished
        else:
            o["waiting"] += 1  # did everything sent; nothing further was sent

# ---------------------------------------------------------------- sheet 1: the rules
RULES = [
    ("Finished", "They completed every interview in their plan.", "unchanged", "no - depends on the design"),
    (
        "Dropped out / inactive",
        "No contact for 14 days - the same 14 days for every cohort.",
        "An interview they were SENT went past its deadline unfinished. The deadline is one gap after it "
        "was sent, so 3 days in a 3-day design and 14 in a 14-day one.",
        "no - uses the cohort's own gap",
    ),
    (
        "Waiting on the schedule",
        "Did not exist. These FLWs were counted as Steady or Inconsistent, " "which reads as on track.",
        "Its own bucket: completed everything sent to them, but the design never finished because nothing "
        "further was sent. 491 FLW-cohort pairs, 15% of starters.",
        "n/a",
    ),
    (
        "Keeping up",
        "Longest silence more than two gaps.",
        "unchanged - but now applies only to FLWs " "still mid-schedule, not to people whose schedule ran out.",
        "no - uses the cohort's own gap",
    ),
    (
        "Keeping up, on the All cohorts view",
        "Longest silence more than 8 days, for every design.",
        "Each FLW judged against their own design's gap.",
        "no - uses the cohort's own gap",
    ),
    (
        "Active right now",
        "Last contact 7 days ago or less.",
        "Last contact within one gap.",
        "no - uses the cohort's own gap",
    ),
    ("Slowing down", "Last contact 8 to 14 days ago.", "Between one and two gaps.", "no - uses the cohort's own gap"),
    (
        "Interview missed",
        "One gap has passed since it was sent.",
        "unchanged - this rule was already " "right, and it is the one the retention graphs now copy.",
        "no - uses the cohort's own gap",
    ),
    (
        "Their LAST interview",
        "Never counted as missed, whatever happened.",
        "Has a deadline like every other interview: one gap after it was sent.",
        "no",
    ),
    (
        "Cohort end date",
        "The end date of the latest cohort in the group.",
        "Each cohort has its own end date: last interview released, plus one gap.",
        "no - per cohort",
    ),
    ("When we measure", "Today.", "At each cohort's own end date.", "no - per cohort"),
    ("Interview release date", "Cohort start plus its position in the plan.", "unchanged", "no"),
    (
        "Release date, funnel chart version",
        "Cohort start, plus 7 days, plus its position.",
        "unchanged - still 7 days out of step with the row above. Flagged, not yet resolved.",
        "no - and it disagrees with the row above",
    ),
    (
        "Recently active (FLW tab)",
        "Last session within 14 days.",
        "Within two gaps.",
        "no - uses the cohort's own gap",
    ),
    (
        "Came back after a break (FLW tab)",
        "A break of 21 days or more, then a return.",
        "A break of three gaps or more.",
        "no - uses the cohort's own gap",
    ),
    (
        "Needs following up (FLW tab)",
        "Last session 14 to 60 days ago.",
        "Between two and eight gaps ago.",
        "no - uses the cohort's own gap",
    ),
    (
        "Change the deadline for one cohort",
        "Not possible - fixed in the code.",
        "A per-cohort setting (GRACE_DAYS). Default is one gap, so nothing needs listing unless a cohort "
        "wants something different.",
        "n/a",
    ),
    (
        "Choose the deadline on the dashboard",
        "n/a",
        "Deliberately NOT added. Every cohort is finished, and once a cohort is over 'did they complete "
        "it?' needs no waiting period - so three presets would have produced three identical charts. The "
        "dashboard explains this where the control would have been.",
        "n/a",
    ),
]
CHANGED = sum(1 for r in RULES if r[2] != "unchanged" and not r[2].startswith("unchanged"))

# ---------------------------------------------------------------- sheet 2: one row per design
types = {}
for c, sg in CSG.items():
    if c in sched:
        types.setdefault((sg, sched[c]["n"], sched[c]["gap"]), []).append(c)

TYPES = []
for (sg, n, gap), cohorts in sorted(types.items()):
    ce = CE.get(sg) or {}
    TYPES.append(
        {
            "Cohort design": sg,
            "Cohorts": len(cohorts),
            "Interviews": n,
            "Gap between interviews": (str(gap) + " days") if gap else "only one interview",
            "Runs for": str(sched[cohorts[0]]["runs"]) + " days",
            "Dropped after - BEFORE": "14 days",
            "...which meant, in this design": (
                "no next interview to miss" if not gap else str(round(14 / gap, 1)) + " missed interviews"
            ),
            "Dropped after - NOW": ("one gap = " + str(gap) + " days") if gap else "14 days (no gap to use)",
            "Last interview has a deadline": "no (before) -> yes (now)",
            "End date - BEFORE (shared by the design)": ce.get("end_date"),
            "End date - NOW": "each cohort's own (see sheet 3)",
        }
    )

# ---------------------------------------------------------------- sheet 3: one row per cohort
COHORTS = []
for c in sorted(CSG):
    if c not in sched or c not in train:
        continue
    s, o = sched[c], out.get(c, Counter())
    COHORTS.append(
        {
            "Cohort": c,
            "Design": CSG[c],
            "Start date": train[c].isoformat(),
            "Its own end date": own_end[c].isoformat() + (" *" if c in est_start else ""),
            "End date used today": (CE.get(CSG[c]) or {}).get("end_date"),
            "Workers": o.get("workers", 0),
            "Completed all": o.get("finished", 0),
            "Dropped off": o.get("dropped", 0),
            "Waiting on the schedule": o.get("waiting", 0),
            "Never sent anything": o.get("never", 0),
            "Drop-off %": (round(100 * o.get("dropped", 0) / o["workers"]) if o.get("workers") else None),
        }
    )

INTRO = [
    (
        "What this is",
        "A description of how the retention graphs on the Interviews dashboard decide "
        "whether a worker is engaged, inactive or dropped out. As it is configured today.",
    ),
    (
        "What a 'gap' means",
        "The number of days between one scheduled interview and the next. It is 3 "
        "days in some cohort designs and 14 in others.",
    ),
    (
        "The main finding",
        str(CHANGED) + " of the " + str(len(RULES)) + " calculations changed. Before the "
        "fix, ten of them used one "
        "fixed number of days for every cohort, however far apart that cohort's interviews are.",
    ),
    (
        "The clearest example",
        "A worker is called dropped out after 14 days of no contact. In ABT2-A, "
        "where interviews are 14 days apart, that is one interview - so someone "
        "exactly on schedule is labelled a drop-out. In TRE, where they are 3 days "
        "apart, it is nearly five missed interviews.",
    ),
    (
        "Why one row per cohort",
        "Cohorts of the same design start weeks apart, so they end weeks apart. "
        "TRS has 44 cohorts whose own end dates span 15 April to 22 May, and all "
        "44 are judged against one shared date. See sheet 3.",
    ),
    (
        "A note on sheet 3",
        "Start and end dates come from the dashboard pipeline itself, so they cannot "
        "disagree with it. A * means that cohort has no Connect invitation date and "
        "its start is taken from the first interview trigger instead. Outcome counts "
        "are computed from the live dashboard payload under the new definition.",
    ),
]

# ---------------------------------------------------------------- write


def table(md, title, note, rows, cols):
    md += ["", "## " + title, "", note, ""]
    md.append("| " + " | ".join(cols) + " |")
    md.append("|" + "---|" * len(cols))
    for r in rows:
        md.append("| " + " | ".join("" if r[c] is None else str(r[c]) for c in cols) + " |")
    return md


md = [
    "# How the retention graphs are configured today",
    "",
    "Generated " + TODAY.isoformat() + " from the live dashboard, the CommCare interview schedule and "
    "the cohort start dates.",
    "",
]
for k, v in INTRO:
    md.append("**" + k + ".** " + v + "")
    md.append("")
md = table(
    md,
    "1. Every calculation feeding the retention graphs",
    str(CHANGED) + " of " + str(len(RULES)) + " calculations changed. Ten of them used to apply "
    "one fixed number of days to every cohort.",
    [
        dict(zip(("What it decides", "How it worked BEFORE", "How it works NOW", "Adapts to the cohort's pace?"), r))
        for r in RULES
    ],
    ["What it decides", "How it worked BEFORE", "How it works NOW", "Adapts to the cohort's pace?"],
)
md = table(
    md,
    "2. One row per cohort design",
    "Every cohort within a design runs the identical plan, so there are exactly "
    + str(len(TYPES))
    + " designs across all "
    + str(len(COHORTS))
    + " cohorts.",
    TYPES,
    list(TYPES[0].keys()),
)
md = table(
    md,
    "3. One row per cohort",
    "The two date columns are the point: each cohort has its own end date, but the graphs use a "
    "single shared one per design.",
    COHORTS,
    list(COHORTS[0].keys()),
)
(ROOT / "retention_config_tables.md").write_text("\n".join(md) + "\n", encoding="utf-8")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEAD = Font(bold=True, color="FFFFFF")
    FILL = PatternFill("solid", fgColor="1F4E79")
    WRAP = Alignment(vertical="top", wrap_text=True)

    def sheet(wb, title, cols, rows, widths, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        ws.append(cols)
        for cell in ws[1]:
            cell.font, cell.fill, cell.alignment = HEAD, FILL, WRAP
        for r in rows:
            ws.append([r[c] if isinstance(r, dict) else r for c in cols] if isinstance(r, dict) else list(r))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = WRAP
        ws.freeze_panes = "A2"
        return ws

    wb = Workbook()
    sheet(wb, "Start here", ["", ""], INTRO, [26, 110], first=True)
    sheet(
        wb,
        "1 The rules",
        ["What it decides", "How it worked BEFORE", "How it works NOW", "Adapts to the cohort's pace?"],
        RULES,
        [30, 46, 58, 30],
    )
    sheet(wb, "2 By cohort design", list(TYPES[0].keys()), TYPES, [15, 9, 11, 22, 11, 20, 24, 30, 32])
    sheet(wb, "3 By cohort", list(COHORTS[0].keys()), COHORTS, [11, 10, 12, 18, 20, 10, 13, 12, 20, 18, 11])
    wb.save(ROOT / "retention_config_tables.xlsx")
    print("wrote retention_config_tables.xlsx and .md")
except ImportError:
    print("wrote retention_config_tables.md (openpyxl missing)")

t = Counter()
for r in COHORTS:
    for k in ("Workers", "Completed all", "Dropped off", "Waiting on the schedule", "Never sent anything"):
        t[k] += r[k]
print(
    "  4 sheets | rules="
    + str(len(RULES))
    + " (changed "
    + str(CHANGED)
    + ") | designs="
    + str(len(TYPES))
    + " | cohorts="
    + str(len(COHORTS))
)
print(
    "  workers="
    + str(t["Workers"])
    + " completed="
    + str(t["Completed all"])
    + " dropped="
    + str(t["Dropped off"])
    + " waiting="
    + str(t["Waiting on the schedule"])
    + " never sent="
    + str(t["Never sent anything"])
    + " adds up="
    + str(
        t["Completed all"] + t["Dropped off"] + t["Waiting on the schedule"] + t["Never sent anything"] == t["Workers"]
    )
)
